import tkinter

from openpyxl import *
from openpyxl.styles import PatternFill
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
from tkinter.ttk import Progressbar
from PIL import Image, ImageTk
from table_var import *
from time import sleep
from idlelib.tooltip import Hovertip

class VRS:
    def __init__(self, title="", resizable=(False, False), icon=r"resourses/icon.ico"):
        self.root = Tk()
        self.root.configure(bg='#dbdbdb')
        self.root.title(title)
        self.root.resizable(resizable[0], resizable[1])
        if icon:
            self.root.iconbitmap(icon)

        # Объекты виджетов окна
        self.vrs_num = Combobox(self.root, values=("019", "034", "039", "054", "058", "078", "086", "097", "115",
                                                   "116", "138", "156", "173", "193", "194", "215", "234", "240",
                                                   "271", "289", "290", "333", "337", "350", "407", "414", "473",
                                                   "500"), width=7, state="readonly")
        self.vrs_perf = Combobox(self.root, values=("01", "02", "03", "04"), width=7, state="readonly")
        self.air = Combobox(self.root, values=("АИР56", "АИР63", "АИР71", "АИР80", "АИР90", "АИР100", "АИР112",
                                               "АИР132", "АИР160", "АИР180"
                                               #, "АИР225", "АИР250"
                                               ), width=7, state="readonly")
        self.air2 = Combobox(self.root, values=("АИР56", "АИР63", "АИР71", "АИР80", "АИР90", "АИР100", "АИР112",
                                                "АИР132", "АИР160", "АИР180"
                                                #, "АИР225", "АИР250"
                                                ), width=7, state="readonly")
        self.vosk = Combobox(self.root, values=("2.5", "2.8", "3.15", "3.55", "4.0", "4.5", "5.0", "5.6"
                                                #"6.3", "7.1", "8.0", "9.0", "10.0", "11.2", "12.5"
                                                ), width=7, state="readonly")
        self.vosk2 = Combobox(self.root, values=("2.5", "2.8", "3.15", "3.55", "4.0", "4.5", "5.0", "5.6"
                                                 #, "6.3", "7.1", "8.0", "9.0", "10.0", "11.2", "12.5"
                                                 ), width=7, state="readonly")
        self.vnv5012_widht = Combobox(self.root, values=("160", "180"), width=7, state="readonly")
       #self.vnv4816_widht = Combobox(self.root, values=("240", "320", "400"), width=7, state="readonly")
        self.vov5012_widht = Combobox(self.root, values=("180", "220", "260", "310"), width=7, state="readonly")
       #self.vov4816_widht = Combobox(self.root, values=("240", "320", "400"), width=7, state="readonly")
        self.filterbox = IntVar()   # заводим int переменную
        self.ventbox = IntVar()
        self.vent2box = IntVar()
        self.vnv5012box = IntVar()
        #self.vnv4816box = IntVar()
        self.vov5012box = IntVar()
        #self.vov4816box = IntVar()
        self.ekobox = IntVar()
        self.vertklapbox = IntVar()
        self.ppbox = IntVar()
        self.promkambox = IntVar()
        count = ([i for i in range(1,11)])
        self.entryfilter = Spinbox(self.root, values=(count), bd=2)  # заводим спинбокс
        self.entryvent = Spinbox(self.root, values=(count), width=7, bd=2)
        self.entryvent2 = Spinbox(self.root, values=(count), width=7, bd=2)
        self.entryvnv5012 = Spinbox(self.root, values=(count), width=7, bd=2)
       #self.entryvnv4816 = Spinbox(self.root, values=(count), width=7, bd=2)
        self.entryvov5012 = Spinbox(self.root, values=(count), width=7, bd=2)
       #self.entryvov4816 = Spinbox(self.root, values=(count), width=7, bd=2)
        self.entryeko = Spinbox(self.root, values=(count), bd=2)
        self.entryvertklap = Spinbox(self.root, values=(count), bd=2)
        self.entrypp = Spinbox(self.root, values=(count), bd=2)
        self.entrypromkam = Spinbox(self.root, values=(count), bd=2)
        self.pb = Progressbar(self.root, orient=HORIZONTAL, mode="determinate", length=400)

    # Запуск окна программы
    def run(self):
        self.draw_widgets()
        self.root.mainloop()
    # Построение окна программы
    def draw_widgets(self):
        deFont = ("Roboto", 10, "bold")
        bgColor = '#dbdbdb'
        # Вставляем лого
        self.logo_upload = Image.open('resourses/upload.png')
        self.logo_close = Image.open('resourses/close.png')
        self.logo_question = Image.open('resourses/question.png')
        self.logo = Image.open('resourses/logo.png')
        self.logo_upload = self.logo_upload.resize((242, 42), Image.Resampling.LANCZOS)
        self.logo_close = self.logo_close.resize((230, 42), Image.Resampling.LANCZOS)
        self.logo_question = self.logo_question.resize((27, 27), Image.Resampling.LANCZOS)
        self.logo = self.logo.resize((120, 54), Image.Resampling.LANCZOS)
        self.logo_upload = ImageTk.PhotoImage(self.logo_upload)
        self.logo_close = ImageTk.PhotoImage(self.logo_close)
        self.logo_question = ImageTk.PhotoImage(self.logo_question)
        self.logo = ImageTk.PhotoImage(self.logo)
        self.logo_label_upload = Label(image=self.logo_upload)
        self.logo_label_close = Label(image=self.logo_close)
        self.logo_label_question = Label(image=self.logo_question)
        self.logo_label = Label(image=self.logo)
        self.logo_label_upload.image = self.logo_upload
        self.logo_label_close.image = self.logo_close
        self.logo_label_question.image = self.logo_question
        self.logo_label.image = self.logo

        self.logo_label.grid(row=1, column=2, columnspan=2, rowspan=2)
        Label(self.root, text="Выберите типоразмер VRS:", justify=LEFT, bg = bgColor, font=deFont).grid(
                                                                        row=1, column=0, sticky=W)
        self.vrs_num.grid(row=1, column=1, sticky=W + E, padx=5, pady=8)
        self.vrs_num.current(0)   # устанавливаем дефолтную позицию выпадающего меню
        Label(self.root, text="Выберите исполнение VRS:", justify=LEFT, bg = bgColor, font=deFont).grid(row=2,
                                                                                                         column=0,
                                                                                                     sticky=W)
        self.vrs_perf.grid(row=2, column=1, sticky=W + E, padx=5, pady=8)       # вывод выпадающего меню
        self.vrs_perf.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        Label(self.root, text="Выберите параметры блоков:", justify=LEFT, bg = bgColor, font=("Roboto", 12,
                                                                        "bold")).grid(row=3, column=0, sticky=W) # вывод строки с текстом
        self.gqbutton = Button(self.root, width=25, height=25, bg = bgColor, image=self.logo_question, border=0,
               command=self.info)
        self.gqbutton.grid(row=3, column=1, sticky=W, padx=5)
        Hovertip(self.gqbutton, 'Информация', hover_delay=400)
        Checkbutton(self.root, text="Блок фильтра", bg = bgColor, justify=LEFT, font=deFont,
                    variable=self.filterbox).grid(row=4, column=0, sticky=W)         # строка чекбокса
        self.entryfilter.grid(row=4, column=1, columnspan=3, sticky=W+E, padx=5, pady=8)  # рисуем спинбокс
        Checkbutton(self.root, text="Блок вентилятора ВОСК", font=deFont, justify=LEFT, bg = bgColor,
                                                        variable=self.ventbox).grid(row=5, column=0, sticky=W)   #строка чекбокса
        self.vosk.grid(row=5, column=1, sticky=W, padx=5, pady=8)
        self.vosk.current(0)
        self.air.grid(row=5, column=2, sticky=W, padx=5, pady=8)           # выпадающая менюшка аиров
        self.air.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        self.entryvent.grid(row=5, column=3, sticky=W, padx=5, pady=8)
        Checkbutton(self.root, text="Блок вентилятора ВОСК 2", justify=LEFT, bg = bgColor, font=deFont, \
                                                                                    variable=self.vent2box).grid(
                                                                    row=6, column=0, sticky=W)  # строка чекбокса
        self.vosk2.grid(row=6, column=1, sticky=W, padx=5, pady=8)
        self.vosk2.current(0)
        self.air2.grid(row=6, column=2, sticky=W, padx=5, pady=8)  # выпадающая менюшка аиров
        self.air2.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        self.entryvent2.grid(row=6, column=3, sticky=W, padx=5, pady=8)
        Checkbutton(self.root, text="Блок ВНВ 5012", justify=LEFT, bg = bgColor, font=deFont,
                    variable=self.vnv5012box).grid(row=7, column=0, sticky=W)        # строка чекбокса
        self.vnvqbutton = Button(self.root, width=25, height=25, bg = bgColor, image=self.logo_question, border=0,
               command=self.vnvinfo)
        Hovertip(self.vnvqbutton, 'Информация', hover_delay=400)
        self.vnvqbutton.grid(row=7, column=0, sticky=E, padx=5)
        self.vnv5012_widht.grid(row=7, column=1, sticky=W, padx=5, pady=8)  # выпадающая менюшка внв
        self.vnv5012_widht.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        self.entryvnv5012.grid(row=7, column=2, columnspan=3, sticky=W+E, padx=5, pady=8)
        # Checkbutton(self.root, text="Блок ВНВ 4816", justify=LEFT, variable=self.vnv4816box).grid(row=6, column=0, sticky=W)      # строка чекбокса
        # self.vnv4816_widht.grid(row=6, column=1, sticky=W)  # выпадающая менюшка внв
        # self.vnv4816_widht.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        # self.entryvnv4816.grid(row=6, column=2, columnspan=2, sticky=W)
        Checkbutton(self.root, text="Блок ВОВ 5012", justify=LEFT, bg = bgColor, font=deFont,
                    variable=self.vov5012box).grid(row=8, column=0, sticky=W)     # строка чекбокса
        self.vovqbutton = Button(self.root, width=25, height=25, bg = bgColor, image=self.logo_question, border=0,
                                 command=self.vovinfo)
        self.vovqbutton.grid(row=8, column=0, sticky=E, padx=5)
        Hovertip(self.vovqbutton, 'Информация', hover_delay=400)
        self.vov5012_widht.grid(row=8, column=1, sticky=W, padx=5, pady=8)  # выпадающая менюшка внв
        self.vov5012_widht.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        self.entryvov5012.grid(row=8, column=2, columnspan=3, sticky=W+E, padx=5, pady=8)
        # Checkbutton(self.root, text="Блок ВОВ 4816", justify=LEFT, variable=self.vov4816box).grid(row=8, column=0, sticky=W)       # строка чекбокса
        # self.vov4816_widht.grid(row=8, column=1, sticky=W)  # выпадающая менюшка внв
        # self.vov4816_widht.current(0)  # устанавливаем дефолтную позицию выпадающего меню
        # self.entryvov4816.grid(row=8, column=2, columnspan=2, sticky=W)
        Checkbutton(self.root, text="Блок ЭКО", justify=LEFT, bg = bgColor, font=deFont, variable=self.ekobox).grid(
            row=9, column=0, sticky=W)  # строка чекбокса
        self.entryeko.grid(row=9, column=1, columnspan=3, sticky=W+E, padx=5, pady=8)
        Checkbutton(self.root, text="Блок вертикального клапана", justify=LEFT, bg = bgColor, font=deFont,
                    variable=self.vertklapbox).grid(row=10, column=0, sticky=W)  # строка чекбокса
        self.entryvertklap.grid(row=10, column=1, columnspan=3, sticky=W+E, padx=5, pady=8)
        Checkbutton(self.root, text="Блок пластинчатого утилизатора", justify=LEFT, bg = bgColor, font=deFont,
                    variable=self.ppbox).grid(row=11, column=0, sticky=W)  # строка чекбокса
        self.entrypp.grid(row=11, column=1, columnspan=3, sticky=W+E, padx=5, pady=8)
        Checkbutton(self.root, text="Блок камеры промежуточной,\nповорот вверх(вниз)",justify=LEFT, bg = bgColor,
                    font=deFont, variable=self.promkambox).grid(row=12, column=0, sticky=W)  # строка чекбокса
        self.entrypromkam.grid(row=12, column=1, columnspan=3, sticky=W+E, padx=5, pady=8) # рисуем спинбокс
        # Кнопки выгрузки и закрытия
        Button(self.root, width=230, height=40, image=self.logo_upload, pady=20, command=self.action).grid(row=13,
                                                                                      column=0, sticky=S)  # Кнопка выгрузки
        Button(self.root, width=207, height=40, image=self.logo_close, padx=10, pady=20,
               command=self.root.destroy).grid(row=13, column=1, columnspan=3, sticky=W)  # Кнопка закрытия
        self.pb.grid(row=14, columnspan=3)
        Label(self.root, bg = bgColor, text="v0.1.2b", justify=LEFT).grid(row=14, column=3, sticky=E)
    # Конопка INFO
    def info(self, title="INFO", resizable=(False, False), icon=r"resourses/info.ico"):
        draw = tkinter.Toplevel()
        draw.title(title)
        draw.resizable(resizable[0], resizable[1])
        draw.configure(bg='#dbdbdb')
        if icon:
            draw.iconbitmap(icon)

        Label(draw, text="- Для блоков ВОСК выбираем типоразмер ВОСК, типоразмер двигателя и "
                                             "количество.\n- Для блока ВНВ и ВОВ "
                                             "выбираем ширину ТО и количество.\n- Для всего остального на выбор только "
                                             "количество.", justify=LEFT, padx=15, pady=15, bg='#dbdbdb', font=("", 10,
                                                                              "bold")).grid(row=0, column=0, sticky=W)
        Button(draw, width=10, height=2, text="Ок", font=("", 10), command=draw.destroy).grid(row=1, column=0)
    # Кнопка INFO для ВНВ
    def vnvinfo(self, title="INFO", resizable=(False, False), icon=r"resourses/info.ico"):
        draw = tkinter.Toplevel()
        draw.title(title)
        draw.resizable(resizable[0], resizable[1])
        draw.configure(bg='#dbdbdb')
        if icon:
            draw.iconbitmap(icon)

        Label(draw, text="160 мм если диаметр коллектора <76 мм\n180 мм если диаметр коллектора =76 мм",
              justify=LEFT, padx=15, pady=15, bg='#dbdbdb', font=("", 10, "bold")).grid(row=0, column=0, sticky=W)
        Button(draw, width=8, height=2, text="Ок", font=("", 10), command=draw.destroy).grid(row=1, column=0)
    # Кнопка INFO для ВОВ
    def vovinfo(self, title="INFO", resizable=(False, False), icon=r"resourses/info.ico"):
        draw = tkinter.Toplevel()
        draw.title(title)
        draw.resizable(resizable[0], resizable[1])
        draw.configure(bg='#dbdbdb')
        if icon:
            draw.iconbitmap(icon)

        Label(draw, text="Для теплообменников с рядностью:\n1-6 рядов - 180 мм\n7-8 рядов - 220 мм\n9-10 рядов - 260 "
                         "мм\n11-12 рядов - 310 мм",
              justify=LEFT, padx=15, pady=15, bg='#dbdbdb', font=("", 10, "bold")).grid(row=0, column=0, sticky=W)
        Button(draw, width=8, height=2, text="Ок", font=("", 10), command=draw.destroy).grid(row=1, column=0)

    # Вывод данных в excel
    def action(self):
        # Открываем excel документ
        wb = Workbook()
        ws = wb.active
        ws.title = 'Перечень'
        global sheet_form
        # Присваиваем переменным значения из полей ввода
        vrs_num = self.vrs_num.get()
        vrs_perf = self.vrs_perf.get()
        vosk = self.vosk.get()
        vosk2 = self.vosk2.get()
        air = self.air.get()
        air2 = self.air2.get()
        vnv5012 = self.vnv5012_widht.get()
        vov5012 = self.vov5012_widht.get()
        #формирование заголовка в выводимой таблице
        ws.cell(row=1, column=1).value = 'Перечень VRS-500-' + vrs_num + '-' + vrs_perf
        ws.cell(row=2, column=1).value = 'Обозначение'
        ws.cell(row=2, column=2).value = 'Наименование'
        ws.cell(row=2, column=3).value = 'Кол-во'
        ws.cell(row=2, column=4).value = 'Материал'
        ws.merge_cells('A1:D1')
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 14
        # Фильтр
        if self.filterbox.get():
            filteramount = int(self.entryfilter.get())
            ws.append(['Блок фильтра', '', filteramount])

            cell = filter_table(vrs_num, vrs_perf)
            for sign, name, amount, material in cell:
                ws.append([sign.value, name.value, int(amount.value*filteramount), material.value])
        self.pb['value'] = 10
        self.pb.update()
        # ВОСК
        if self.ventbox.get():
            ventamount = int(self.entryvent.get())
            ws.append(['Блок вентилятора ВОСК ' + self.vosk.get() + ' ' + self.air.get(), '', ventamount])
            #ws.append(['Разблюдовка:'])

            vent = vent_table(vrs_num, vrs_perf, vosk)
            aircell = air_table(vrs_perf, air, vosk)
            for sign, name, amount, material in vent:
                ws.append([sign.value, name.value, int(amount.value * ventamount), material.value])
            for sign, name, amount, material in aircell:
                ws.append([sign.value, name.value, int(amount.value * ventamount), material.value])
        self.pb['value'] = 20
        self.pb.update()
        # ВОСК2
        if self.vent2box.get():
            ventamount2 = int(self.entryvent2.get())
            ws.append(['Блок вентилятора ВОСК ' + self.vosk2.get() + ' ' + self.air2.get(), '', ventamount2])
            #ws.append(['Разблюдовка:'])

            vent2 = vent_table2(vrs_num, vrs_perf, vosk2)
            aircell2 = air_table2(vrs_perf, air2, vosk2)
            for sign, name, amount, material in vent2:
                ws.append([sign.value, name.value, int(amount.value * ventamount2), material.value])
            for sign, name, amount, material in aircell2:
                ws.append([sign.value, name.value, int(amount.value * ventamount2), material.value])
        self.pb['value'] = 30
        self.pb.update()
        # ВНВ
        if self.vnv5012box.get():
            vnv5012amount = int(self.entryvnv5012.get())
            ws.append(['Блок ВНВ 5012', '', vnv5012amount])

            vnvcell = vnv5012_table(vrs_num, vrs_perf, vnv5012)
            for sign, name, amount, material in vnvcell:
                ws.append([sign.value, name.value, int(amount.value*vnv5012amount), material.value])
        self.pb['value'] = 40
        self.pb.update()
        # ВОВ
        if self.vov5012box.get():
            vov5012amount = int(self.entryvov5012.get())
            ws.append(['Блок ВОВ 5012', '', vov5012amount])

            vovcell = vov5012_table(vrs_num, vrs_perf, vov5012)
            for sign, name, amount, material in vovcell:
                ws.append([sign.value, name.value, int(amount.value * vov5012amount), material.value])
        self.pb['value'] = 50
        self.pb.update()
        # ЭКО
        if self.ekobox.get():
            ekoamount = int(self.entryeko.get())
            ws.append(['Блок ЭКО', '', ekoamount])

            ekocell = eko_table(vrs_num, vrs_perf)
            for sign, name, amount, material in ekocell:
                ws.append([sign.value, name.value, int(amount.value * ekoamount), material.value])
        self.pb['value'] = 60
        self.pb.update()
        # Вертикальный клапан
        if self.vertklapbox.get():
            vertklapamount = int(self.entryvertklap.get())
            ws.append(['Блок вертикального клапана', '', vertklapamount])

            vertklapcell = vertklap_table(vrs_num, vrs_perf)
            for sign, name, amount, material in vertklapcell:
                ws.append([sign.value, name.value, int(amount.value * vertklapamount), material.value])
        self.pb['value'] = 70
        self.pb.update()
        # Пластинчатый утилизатор
        if self.ppbox.get():
             ppamount = int(self.entrypp.get())
             ws.append(['Блок пластинчатого утилизатора', '', ppamount])

             ppcell = pp_table(vrs_num, vrs_perf)
             for sign, name, amount, material in ppcell:
                ws.append([sign.value, name.value, int(amount.value * ppamount), material.value])
        self.pb['value'] = 85
        self.pb.update()
        # Камера промежуточная
        if self.promkambox.get():
             promkamamount = int(self.entrypromkam.get())
             ws.append(['Блок камеры промежуточной', '', promkamamount])

             promkamcell = promkam_table(vrs_num, vrs_perf)
             for sign, name, amount, material in promkamcell:
                 ws.append([sign.value, name.value, int(amount.value * promkamamount), material.value])
        self.pb['value'] = 100
        self.pb.update()
        #закрашиваем ячейки с названиями блоков
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Блок фильтра":
                    filtercellrow = cell.row
                    filtercellcolumn = cell.column
                    ws.cell(row=filtercellrow, column=filtercellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=filtercellrow, column=filtercellcolumn + 1).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                    ws.cell(row=filtercellrow, column=filtercellcolumn + 2).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                    ws.cell(row=filtercellrow, column=filtercellcolumn + 3).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                elif cell.value == 'Блок вентилятора ВОСК ' + self.vosk.get() + ' ' + self.air.get():
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                               fgColor="abaeb3")
                elif cell.value == 'Блок вентилятора ВОСК ' + self.vosk2.get() + ' ' + self.air2.get():
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок ВНВ 5012':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок ВОВ 5012':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок ЭКО':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок вертикального клапана':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок пластинчатого утилизатора':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                elif cell.value == 'Блок камеры промежуточной':
                    voskcellrow = cell.row
                    voskcellcolumn = cell.column
                    ws.cell(row=voskcellrow, column=voskcellcolumn).fill = PatternFill('solid', fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 1).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 2).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")
                    ws.cell(row=voskcellrow, column=voskcellcolumn + 3).fill = PatternFill('solid',
                                                                                           fgColor="abaeb3")

        if self.filterbox.get() == 0 and self.ventbox.get() == 0 and self.vent2box.get() == 0 and \
                self.vnv5012box.get()  == 0 and self.vov5012box.get() == 0 and self.ekobox.get() == 0 and \
                self.vertklapbox.get() == 0 and self.ppbox.get() == 0 and self.promkambox.get() == 0:
            messagebox.showerror("VRS", "Выберите хотя бы 1 блок")
        else:
            wb.save('Перечень VRS-500-' + vrs_num + '-' + vrs_perf + '.xlsx')
            messagebox.showinfo("VRS", 'Перечень VRS-500-' + vrs_num + '-' + vrs_perf + '\nвыгружен в корень папки')


if __name__ == "__main__":
    program = VRS("Перечень деталей")
    program.run()